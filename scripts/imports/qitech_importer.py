from __future__ import annotations

import argparse
import hashlib
import os
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


POSITION_SECTIONS = {"SRP", "DIR", "MEZAN", "NTN-B", "OUTROSFUNDOS"}

DRE_CATEGORIES = {
    "taxa_gestao": "Taxa de Gestao",
    "taxa_administracao": "Taxa de Administracao",
    "taxa_custodia": "Taxa de Custodia",
    "auditoria": "Auditoria",
    "servicos_cobranca": "Servicos de Cobranca",
    "iof": "IOF",
    "cetip": "CETIP",
    "selic": "SELIC",
    "consultoria": "Consultoria",
    "rating": "Rating",
    "outras_despesas": "Outras Despesas",
    "pdd": "PDD - Provisao para Devedores Duvidosos",
}


@dataclass
class FundQuoteData:
    fund_name: str
    position_date: date
    net_asset_value: Decimal
    quota_value: Decimal
    shares_quantity: Decimal
    daily_return: Decimal
    month_return: Decimal
    year_return: Decimal


@dataclass
class FinancialPositionData:
    asset_class: str
    position_date: date
    code: str
    asset_name: str
    quantity: Decimal
    market_unit_price: Decimal
    gross_value: Decimal
    net_value: Decimal
    indexer: str | None
    maturity_date: date | None


@dataclass
class DreEntryData:
    reference_date: date
    description: str
    amount: Decimal
    translated_history: str
    category: str


@dataclass
class CashImportData:
    header_date: date | None
    treasury_balance: Decimal | None


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip().lower()


def normalized_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_text(value))


def parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip()
    if not text:
        return None
    text = text.replace("%", "").replace("R$", "").strip()
    text = re.sub(r"[^0-9,.\-]", "", text)
    if not text or text in {"-", ".", ","}:
        return None

    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def first_date_in_text(value: Any) -> date | None:
    if value is None:
        return None
    match = re.search(r"\d{2}/\d{2}/\d{4}", str(value))
    if not match:
        return parse_date(value)
    return parse_date(match.group(0))


def first_decimal(values: Iterable[Any]) -> Decimal | None:
    for value in values:
        number = parse_decimal(value)
        if number is not None:
            return number
    return None


def find_header(headers: list[Any], *candidates: str) -> int | None:
    normalized_headers = [normalized_key(header) for header in headers]
    normalized_candidates = [normalized_key(candidate) for candidate in candidates]
    for candidate in normalized_candidates:
        for index, header in enumerate(normalized_headers):
            if candidate == header or candidate in header:
                return index
    return None


def row_value(row: tuple[Any, ...], headers: list[Any], *candidates: str) -> Any:
    index = find_header(headers, *candidates)
    if index is None or index >= len(row):
        return None
    return row[index]


def iter_rows(ws: Worksheet) -> list[tuple[Any, ...]]:
    return [tuple(row) for row in ws.iter_rows(values_only=True)]


def extract_header_data(workbook: Any) -> tuple[str, date]:
    fund_name = ""
    position_date = None

    for max_rows in (10, 20):
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(max_row=max_rows):
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value)
                    normalized = normalize_text(text)

                    if "cliente:" in normalized and not fund_name:
                        fund_name = text.split(":", 1)[1].strip()

                    if "data de posicao" in normalized:
                        match = re.search(r"\d{2}/\d{2}/\d{4}", text)
                        if match:
                            position_date = parse_date(match.group(0))
                            print(
                                "Data de posicao encontrada "
                                f"na aba '{worksheet.title}', celula {cell.coordinate}: {text}"
                            )
                            break
                if position_date is not None:
                    break
            if position_date is not None:
                break
        if position_date is not None:
            break

    if not fund_name:
        raise ValueError("Nao foi possivel encontrar o cabecalho CLIENTE no arquivo de carteira.")
    if position_date is None:
        raise ValueError("Nao foi possivel encontrar a Data de posicao no arquivo de carteira.")

    return fund_name, position_date


def extract_rentability(rows: list[tuple[Any, ...]], fund_name: str, position_date: date) -> FundQuoteData:
    header_index = None
    headers: list[Any] = []

    for index, row in enumerate(rows):
        if any(normalized_key(cell) == "codindexador" for cell in row):
            header_index = index
            headers = list(row)
            break

    if header_index is None:
        raise ValueError("Secao Rentabilidade nao encontrada: coluna 'Cod. Indexador' ausente.")

    code_index = find_header(headers, "Cod. Indexador")
    if code_index is None:
        raise ValueError("Coluna 'Cod. Indexador' nao encontrada na secao Rentabilidade.")

    rentability_rows: dict[str, tuple[Any, ...]] = {}
    for row in rows[header_index + 1 :]:
        code = normalize_text(row[code_index] if code_index < len(row) else None)
        if code in {"patrimon", "cota", "vlr cota", "qtd cota"}:
            rentability_rows[code] = row
        if len(rentability_rows) == 4:
            break

    try:
        patrimon = rentability_rows["patrimon"]
        cota = rentability_rows["cota"]
        vlr_cota = rentability_rows["vlr cota"]
        qtd_cota = rentability_rows["qtd cota"]
    except KeyError as error:
        raise ValueError(f"Linha obrigatoria ausente na secao Rentabilidade: {error.args[0]}") from error

    net_asset_value = parse_decimal(row_value(patrimon, headers, "Valor Patrimonio", "Valor Patrimônio"))
    quota_value = first_decimal(vlr_cota[code_index + 1 :])
    shares_quantity = first_decimal(qtd_cota[code_index + 1 :])
    daily_return = parse_decimal(row_value(cota, headers, "(%) Variacao Diaria", "(%) Variação Diaria"))
    month_return = parse_decimal(row_value(cota, headers, "(%) Variacao Mensal", "(%) Variação Mensal"))
    year_return = parse_decimal(row_value(cota, headers, "(%) Variacao Anual", "(%) Variação Anual"))

    required = {
        "Valor Patrimonio": net_asset_value,
        "Vlr Cota": quota_value,
        "Qtd Cota": shares_quantity,
        "Variacao Diaria": daily_return,
        "Variacao Mensal": month_return,
        "Variacao Anual": year_return,
    }
    missing = [name for name, value in required.items() if value is None]
    if missing:
        raise ValueError(f"Campos ausentes na Rentabilidade: {', '.join(missing)}")

    return FundQuoteData(
        fund_name=fund_name,
        position_date=position_date,
        net_asset_value=net_asset_value,
        quota_value=quota_value,
        shares_quantity=shares_quantity,
        daily_return=daily_return,
        month_return=month_return,
        year_return=year_return,
    )


def classify_asset_class(section: str, asset_name: str) -> str:
    normalized_section = normalized_key(section)
    normalized_asset_name = normalize_text(asset_name)

    if normalized_section == "outrosfundos":
        if (
            "a vencer" in normalized_asset_name
            or "vencidos" in normalized_asset_name
            or "bristol" in normalized_asset_name
        ):
            return "direitos_creditorios"

        return "outros_fundos"

    if normalized_section == "srp":
        return "senior"

    if normalized_section == "mezan":
        return "mezanino"

    if normalized_section == "ntnb":
        return "ntnb"

    if normalized_section == "dir":
        return "direitos_creditorios"

    if normalized_section == "pdddir":
        return "pdd"

    return normalized_section or section


def classify_outros_fundos(code: str, fund_name: str) -> str:
    normalized_code = normalize_text(code)
    normalized_fund_name = normalize_text(fund_name)

    if "bris" in normalized_code or "bristol" in normalized_fund_name:
        return "direitos_creditorios"

    return "outros_fundos"


def extract_positions(rows: list[tuple[Any, ...]], default_position_date: date) -> list[FinancialPositionData]:
    positions: list[FinancialPositionData] = []

    for index, row in enumerate(rows):
        section = None
        for cell in row:
            key = normalized_key(cell)
            if key in {normalized_key(item) for item in POSITION_SECTIONS}:
                section = str(cell).strip()
                break
        if section is None:
            continue

        header_index = None
        for candidate_index in range(index + 1, min(index + 6, len(rows))):
            candidate = rows[candidate_index]
            if find_header(list(candidate), "Nome Papel", "Valor Bruto", "Valor Liquido") is not None:
                header_index = candidate_index
                break
        if header_index is None:
            continue

        headers = list(rows[header_index])
        for data_row in rows[header_index + 1 :]:
            if all(cell is None or str(cell).strip() == "" for cell in data_row):
                break
            first_cell = normalize_text(data_row[0] if data_row else None)
            if normalized_key(first_cell) in {normalized_key(item) for item in POSITION_SECTIONS}:
                break

            is_outros_fundos = normalized_key(section) == "outrosfundos"
            asset_name = row_value(data_row, headers, "Fundo", "Nome Papel", "Papel", "Ativo")
            gross_value = (
                parse_decimal(row_value(data_row, headers, "Valor Liquido", "Valor Líquido"))
                if is_outros_fundos
                else parse_decimal(row_value(data_row, headers, "Valor Bruto"))
            )
            net_value = parse_decimal(row_value(data_row, headers, "Valor Liquido", "Valor Líquido"))
            if not asset_name or (gross_value is None and net_value is None):
                continue

            quantity = parse_decimal(row_value(data_row, headers, "Quantidade", "Qtd")) or Decimal("0")
            market_unit_price = parse_decimal(row_value(data_row, headers, "PU Mercado", "PU")) or Decimal("0")
            position_date = parse_date(row_value(data_row, headers, "Data Posicao", "Data Posição")) or default_position_date
            code = str(row_value(data_row, headers, "Codigo", "Código", "Cod.") or "").strip()
            asset_class = (
                classify_outros_fundos(code, str(asset_name))
                if is_outros_fundos
                else classify_asset_class(section, str(asset_name))
            )

            if is_outros_fundos:
                print(
                    "OutrosFundos encontrado: "
                    f"codigo='{code}', nome='{asset_name}', valor_liquido={net_value or gross_value}, "
                    f"asset_class={asset_class}"
                )

            positions.append(
                FinancialPositionData(
                    asset_class=asset_class,
                    position_date=position_date,
                    code=code,
                    asset_name=str(asset_name).strip(),
                    quantity=quantity,
                    market_unit_price=market_unit_price,
                    gross_value=gross_value or Decimal("0"),
                    net_value=net_value or gross_value or Decimal("0"),
                    indexer=(str(row_value(data_row, headers, "Indexador") or "").strip() or None),
                    maturity_date=parse_date(row_value(data_row, headers, "Vencimento")),
                )
            )

    return positions


def extract_pdd_positions(rows: list[tuple[Any, ...]], default_position_date: date) -> list[FinancialPositionData]:
    positions: list[FinancialPositionData] = []

    for index, row in enumerate(rows):
        if not any(normalize_text(cell) == "outrosativos" for cell in row):
            continue

        header_index = index + 1
        if header_index >= len(rows):
            continue

        headers = list(rows[header_index])
        date_index = find_header(headers, "Data")
        code_index = find_header(headers, "Codigo", "Código")
        description_index = find_header(headers, "Descricao", "Descrição")
        value_index = find_header(headers, "Valor Total")

        if code_index is None or description_index is None or value_index is None:
            print("Aviso: secao OutrosAtivos encontrada, mas colunas de PDD ausentes.")
            continue

        for data_row in rows[header_index + 1 :]:
            first_value = first_filled_value(data_row)
            if first_value is None:
                continue
            if normalize_text(first_value).startswith("totais"):
                break

            code = str(data_row[code_index] or "").strip()
            description = str(data_row[description_index] or "").strip()
            if normalize_text(code) != "pdd" and "pdd" not in normalize_text(description):
                continue

            total_value = parse_decimal(data_row[value_index])
            if total_value is None or total_value == Decimal("0"):
                continue

            position_date = (
                parse_date(data_row[date_index])
                if date_index is not None and date_index < len(data_row)
                else None
            ) or default_position_date

            print(
                "PDD encontrada em OutrosAtivos: "
                f"data={position_date}, codigo='{code}', descricao='{description}', valor={total_value}"
            )

            positions.append(
                FinancialPositionData(
                    asset_class="pdd",
                    position_date=position_date,
                    code=code,
                    asset_name=description,
                    quantity=Decimal("0"),
                    market_unit_price=Decimal("0"),
                    gross_value=total_value,
                    net_value=total_value,
                    indexer=None,
                    maturity_date=None,
                )
            )

        break

    return positions


def categorize_expense(history: str) -> str:
    normalized = normalize_text(history)
    if "gestao" in normalized:
        return "taxa_gestao"
    if "administracao" in normalized:
        return "taxa_administracao"
    if "custodia" in normalized:
        return "taxa_custodia"
    if "auditoria" in normalized:
        return "auditoria"
    if "cobranca" in normalized:
        return "servicos_cobranca"
    if "iof" in normalized:
        return "iof"
    if "cetip" in normalized:
        return "cetip"
    if "selic" in normalized:
        return "selic"
    if "consultoria" in normalized:
        return "consultoria"
    if "rating" in normalized:
        return "rating"
    return "outras_despesas"


def first_filled_value(row: Iterable[Any]) -> Any:
    for value in row:
        if value is not None and str(value).strip() != "":
            return value
    return None


def extract_caixa_header_date(rows: list[tuple[Any, ...]]) -> date | None:
    for row in rows[:4]:
        for cell in row:
            parsed_date = first_date_in_text(cell)
            if parsed_date is not None:
                print(f"Data do demonstrativo de caixa encontrada no cabecalho: {parsed_date.strftime('%d/%m/%Y')}")
                return parsed_date
    return None


def extract_carteira_cpr_entries(workbook: Any, reference_date: date) -> list[DreEntryData]:
    entries: list[DreEntryData] = []

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            cpr_cell = next((cell for cell in row if normalize_text(cell.value) == "cpr"), None)
            if cpr_cell is None:
                continue

            print(f"Secao CPR da carteira encontrada na aba '{worksheet.title}', celula {cpr_cell.coordinate}.")
            header_cells = next(
                worksheet.iter_rows(min_row=cpr_cell.row + 1, max_row=cpr_cell.row + 1),
                None,
            )
            if header_cells is None:
                continue

            headers = [cell.value for cell in header_cells]
            date_index = find_header(headers, "Data")
            description_index = find_header(headers, "Descricao", "Descrição")
            amount_index = find_header(headers, "Valor")
            history_index = find_header(headers, "Historico Traduzido", "Histórico Traduzido")

            missing_columns = [
                name
                for name, index in {
                    "Descricao": description_index,
                    "Valor": amount_index,
                    "Historico Traduzido": history_index,
                }.items()
                if index is None
            ]
            if missing_columns:
                print(
                    "Aviso: secao CPR ignorada por colunas ausentes "
                    f"na aba '{worksheet.title}': {', '.join(missing_columns)}"
                )
                continue

            for data_cells in worksheet.iter_rows(min_row=cpr_cell.row + 2):
                data_row = tuple(cell.value for cell in data_cells)
                first_value = first_filled_value(data_row)
                if first_value is not None and normalize_text(first_value).startswith("totais"):
                    print(f"Fim da secao CPR da carteira na linha {data_cells[0].row}: {first_value}")
                    break
                if first_value is None:
                    continue

                amount = parse_decimal(data_row[amount_index])
                if amount is None or amount == Decimal("0"):
                    continue

                description = str(data_row[description_index] or "").strip()
                translated_history = str(data_row[history_index] or description).strip()
                entry_date = (
                    parse_date(data_row[date_index])
                    if date_index is not None and date_index < len(data_row)
                    else None
                ) or reference_date
                category = categorize_expense(translated_history)

                print(
                    "Despesa CPR encontrada: "
                    f"aba='{worksheet.title}', linha={data_cells[0].row}, data={entry_date}, "
                    f"descricao='{description}', historico='{translated_history}', "
                    f"valor={amount}, categoria={category}"
                )

                entries.append(
                    DreEntryData(
                        reference_date=entry_date,
                        description=description or translated_history,
                        amount=amount,
                        translated_history=translated_history,
                        category=category,
                    )
                )

            return entries

    if not entries:
        print("Aviso: nenhuma despesa CPR encontrada na carteira.")

    return entries


def extract_treasury_balance(rows: list[tuple[Any, ...]]) -> Decimal | None:
    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            if "saldo em tesouraria" not in normalize_text(value):
                continue

            candidates = list(row[column_index:]) + list(row[: column_index - 1])
            balance = first_decimal(candidates)
            if balance is not None:
                print(
                    "Saldo em Tesouraria encontrado no demonstrativo de caixa: "
                    f"linha={row_index}, coluna={column_index}, valor={balance}"
                )
                return balance

            print(
                "Aviso: linha 'Saldo em Tesouraria' encontrada, "
                f"mas sem valor numerico: linha={row_index}, coluna={column_index}"
            )
            return None

    print("Aviso: Saldo em Tesouraria nao encontrado no demonstrativo de caixa.")
    return None


def extract_section_rows(rows: list[tuple[Any, ...]], section_name: str) -> list[DreEntryData]:
    section_key = normalized_key(section_name)
    entries: list[DreEntryData] = []

    for index, row in enumerate(rows):
        if not any(normalized_key(cell) == section_key for cell in row):
            continue

        header_index = None
        for candidate_index in range(index + 1, min(index + 8, len(rows))):
            candidate = rows[candidate_index]
            if find_header(list(candidate), "Data", "Valor") is not None:
                header_index = candidate_index
                break
        if header_index is None:
            continue

        headers = list(rows[header_index])
        for data_row in rows[header_index + 1 :]:
            if all(cell is None or str(cell).strip() == "" for cell in data_row):
                break
            if data_row and normalized_key(data_row[0]) in {"cpr", "outrosativos"}:
                break

            reference_date = parse_date(row_value(data_row, headers, "Data", "Dt. Movimento", "Data Movimento"))
            amount = parse_decimal(row_value(data_row, headers, "Valor", "Valor Financeiro"))
            description = str(row_value(data_row, headers, "Descricao", "Descrição", "Historico", "Histórico") or "").strip()
            translated_history = str(
                row_value(data_row, headers, "Historico Traduzido", "Histórico Traduzido") or description
            ).strip()

            if reference_date is None or amount is None:
                continue

            category = "pdd" if section_key == "outrosativos" else categorize_expense(translated_history)
            if section_key == "outrosativos" and "pdd" not in normalize_text(translated_history + " " + description):
                continue

            entries.append(
                DreEntryData(
                    reference_date=reference_date,
                    description=description or translated_history,
                    amount=amount,
                    translated_history=translated_history,
                    category=category,
                )
            )

    return entries


def load_carteira(path: Path) -> tuple[FundQuoteData, list[FinancialPositionData], list[DreEntryData]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active
    rows = iter_rows(worksheet)
    fund_name, position_date = extract_header_data(workbook)
    quote = extract_rentability(rows, fund_name, position_date)
    positions = extract_positions(rows, position_date) + extract_pdd_positions(rows, position_date)
    dre_entries = extract_carteira_cpr_entries(workbook, position_date)
    return quote, positions, dre_entries


def load_caixa(path: Path, reference_date: date | None = None) -> CashImportData:
    workbook = load_workbook(path, data_only=True)
    worksheet = (
        workbook["Relatorio_Demonstrativo_Caixa"]
        if "Relatorio_Demonstrativo_Caixa" in workbook.sheetnames
        else workbook.active
    )
    rows = iter_rows(worksheet)
    header_date = extract_caixa_header_date(rows) or reference_date
    treasury_balance = extract_treasury_balance(rows)
    return CashImportData(header_date=header_date, treasury_balance=treasury_balance)


def stable_id(prefix: str, *parts: Any) -> str:
    digest = hashlib.sha1("|".join(str(part) for part in parts).encode("utf-8")).hexdigest()
    return f"{prefix}_{digest[:24]}"


def load_environment() -> None:
    script_dir = Path(__file__).resolve().parent
    root_dir = script_dir.parents[1]
    for env_path in (
        root_dir / ".env",
        root_dir / "apps" / "web" / ".env",
        root_dir / "packages" / "database" / ".env",
    ):
        if env_path.exists():
            load_dotenv(env_path, override=False)


def find_fund_id(cursor: Any, fund_name: str) -> str:
    cursor.execute(
        """
        SELECT id, name
        FROM funds
        WHERE name = %s OR "shortName" = %s
        ORDER BY name
        LIMIT 1
        """,
        (fund_name, fund_name),
    )
    row = cursor.fetchone()
    if row:
        print(f"Fundo encontrado por correspondencia exata: {row[1]} ({row[0]})")
        return row[0]

    cursor.execute(
        """
        SELECT id, name
        FROM funds
        WHERE name ILIKE %s OR "shortName" ILIKE %s
        ORDER BY name
        LIMIT 1
        """,
        (f"%{fund_name}%", f"%{fund_name}%"),
    )
    row = cursor.fetchone()
    if row:
        print(f"Fundo encontrado por busca parcial: {row[1]} ({row[0]})")
        return row[0]

    normalized_name = normalize_text(fund_name)
    words = [word for word in re.split(r"\W+", normalized_name) if len(word) >= 3]
    for word in words:
        cursor.execute(
            """
            SELECT id, name
            FROM funds
            WHERE name ILIKE %s OR "shortName" ILIKE %s
            ORDER BY name
            LIMIT 1
            """,
            (f"%{word}%", f"%{word}%"),
        )
        row = cursor.fetchone()
        if row:
            print(f"Fundo encontrado por palavra-chave '{word}': {row[1]} ({row[0]})")
            return row[0]

    raise ValueError(f"Fundo nao encontrado no banco para o nome extraido: {fund_name}")


def dre_account_ids(cursor: Any) -> dict[str, str]:
    cursor.execute("SELECT id, code FROM dre_accounts")
    rows = cursor.fetchall()
    return {code: account_id for account_id, code in rows}


def upsert_quote(cursor: Any, fund_id: str, quote: FundQuoteData) -> None:
    cursor.execute(
        """
        INSERT INTO fund_quotes (
          id, "fundId", "quoteDate", "quotaValue", "netAssetValue",
          "sharesQuantity", "dailyReturn", "monthReturn", "yearReturn", "createdAt"
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
        ON CONFLICT ("fundId", "quoteDate") DO UPDATE SET
          "quotaValue" = EXCLUDED."quotaValue",
          "netAssetValue" = EXCLUDED."netAssetValue",
          "sharesQuantity" = EXCLUDED."sharesQuantity",
          "dailyReturn" = EXCLUDED."dailyReturn",
          "monthReturn" = EXCLUDED."monthReturn",
          "yearReturn" = EXCLUDED."yearReturn"
        """,
        (
            stable_id("quote", fund_id, quote.position_date),
            fund_id,
            quote.position_date,
            quote.quota_value,
            quote.net_asset_value,
            quote.shares_quantity,
            quote.daily_return,
            quote.month_return,
            quote.year_return,
        ),
    )


def upsert_dre_entries(cursor: Any, fund_id: str, accounts: dict[str, str], entries: list[DreEntryData]) -> int:
    imported_count = 0

    for entry in entries:
        account_id = accounts.get(entry.category)
        if account_id is None:
            print(
                f"Aviso: conta DRE '{entry.category}' nao encontrada. "
                "Usando fallback 'outras_despesas'."
            )
            account_id = accounts.get("outras_despesas")

        if account_id is None:
            print(
                "Aviso: conta DRE fallback 'outras_despesas' nao existe. "
                f"Registro ignorado: {entry.description}"
            )
            continue

        entry_id = stable_id(
            "dre",
            fund_id,
            entry.category,
            entry.reference_date,
            entry.amount,
            entry.description,
            "QITECH",
        )
        cursor.execute(
            """
            INSERT INTO dre_entries (
              id, "fundId", "accountId", "referenceDate", amount, description, source
            )
            VALUES (%s, %s, %s, %s, %s, %s, 'QITECH')
            ON CONFLICT (id) DO UPDATE SET
              "accountId" = EXCLUDED."accountId",
              "referenceDate" = EXCLUDED."referenceDate",
              amount = EXCLUDED.amount,
              description = EXCLUDED.description,
              source = EXCLUDED.source
            """,
            (
                entry_id,
                fund_id,
                account_id,
                entry.reference_date,
                entry.amount,
                entry.description,
            ),
        )
        imported_count += 1

    return imported_count


def upsert_positions(cursor: Any, fund_id: str, positions: list[FinancialPositionData]) -> int:
    for position in positions:
        position_id = stable_id(
            "pos",
            fund_id,
            position.position_date,
            position.asset_class,
            position.code,
            position.asset_name,
        )
        cursor.execute(
            """
            INSERT INTO financial_positions (
              id, "fundId", "positionDate", "assetClass", "assetName",
              quantity, "grossValue", "netValue"
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (id) DO UPDATE SET
              "positionDate" = EXCLUDED."positionDate",
              "assetClass" = EXCLUDED."assetClass",
              "assetName" = EXCLUDED."assetName",
              quantity = EXCLUDED.quantity,
              "grossValue" = EXCLUDED."grossValue",
              "netValue" = EXCLUDED."netValue"
            """,
            (
                position_id,
                fund_id,
                position.position_date,
                position.asset_class,
                position.asset_name,
                position.quantity,
                position.gross_value,
                position.net_value,
            ),
        )
    return len(positions)


def run_import(carteira_path: Path, caixa_path: Path) -> None:
    if not carteira_path.exists():
        raise FileNotFoundError(f"Arquivo de carteira nao encontrado: {carteira_path}")
    if not caixa_path.exists():
        raise FileNotFoundError(f"Arquivo de demonstrativo de caixa nao encontrado: {caixa_path}")

    load_environment()
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_URL nao encontrada. Configure o arquivo .env antes de importar.")

    quote, positions, dre_entries = load_carteira(carteira_path)
    cash_import = load_caixa(caixa_path, quote.position_date)

    with psycopg2.connect(database_url) as connection:
        with connection.cursor() as cursor:
            fund_id = find_fund_id(cursor, quote.fund_name)
            accounts = dre_account_ids(cursor)
            upsert_quote(cursor, fund_id, quote)
            dre_count = upsert_dre_entries(cursor, fund_id, accounts, dre_entries)
            position_count = upsert_positions(cursor, fund_id, positions)
        connection.commit()

    print("Importacao QITECH concluida.")
    print(f"Fundo: {quote.fund_name}")
    print(f"Data de posicao: {quote.position_date.strftime('%d/%m/%Y')}")
    print("fund_quotes: 1 registro inserido/atualizado")
    print(f"dre_entries: {dre_count} registros inseridos/atualizados")
    print(f"financial_positions: {position_count} registros inseridos/atualizados")
    if cash_import.treasury_balance is not None:
        print(f"saldo_tesouraria_caixa: {cash_import.treasury_balance}")


def debug_caixa(filepath: str) -> None:
    workbook = load_workbook(filepath, data_only=True)
    print("Abas disponiveis:")
    for index, sheet_name in enumerate(workbook.sheetnames):
        print(f"{index}: {sheet_name}")

    worksheet = workbook.worksheets[0]
    print(f"\nDiagnostico das primeiras 100 linhas na primeira aba: {worksheet.title}")

    for row in worksheet.iter_rows(max_row=100):
        for cell in row:
            if cell.value is None or str(cell.value).strip() == "":
                continue

            text = str(cell.value)
            normalized = normalize_text(text)
            marker = ""
            if normalized == "cpr":
                marker = " >>> CPR <<<"
            elif "totais" in normalized:
                marker = " >>> TOTAIS <<<"

            print(f"Linha {cell.row}, Col {cell.column}: {text}{marker}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Importa arquivos XLSX QITECH para o banco OSHER.")
    parser.add_argument("carteira", help="Caminho do arquivo ATIVO_CARTEIRA_DIARIA_*.xlsx")
    parser.add_argument("caixa", help="Caminho do arquivo ATIVO_DEMONSTRATIVO_CAIXA_*.xlsx")
    args = parser.parse_args()

    try:
        run_import(Path(args.carteira), Path(args.caixa))
        return 0
    except Exception as error:
        print(f"Erro na importacao QITECH: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    if len(sys.argv) == 3 and sys.argv[1] == "--debug":
        debug_caixa(sys.argv[2])
        sys.exit(0)

    raise SystemExit(main())
